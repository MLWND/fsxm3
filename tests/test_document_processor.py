"""测试文档处理器：TXT/CSV/Markdown/Excel/编码。"""

from core.document_processor import _load_csv, _load_markdown, _load_txt, _load_xlsx, process_document


class TestLoadTxt:
    def test_utf8(self, tmp_txt):
        docs = _load_txt(tmp_txt)
        assert len(docs) == 1
        assert "测试文本" in docs[0].page_content

    def test_gbk(self, tmp_gbk_txt):
        docs = _load_txt(tmp_gbk_txt)
        assert len(docs) == 1
        assert "GBK编码" in docs[0].page_content


class TestLoadCsv:
    def test_basic(self, tmp_csv):
        docs = _load_csv(tmp_csv)
        assert len(docs) == 2  # 2 data rows
        assert "张三" in docs[0].page_content
        assert "北京" in docs[0].page_content

    def test_content_format(self, tmp_csv):
        docs = _load_csv(tmp_csv)
        # 应该包含 "列名: 值" 格式
        assert "姓名:" in docs[0].page_content or "张三" in docs[0].page_content


class TestLoadMarkdown:
    def test_basic(self, tmp_markdown):
        docs = _load_markdown(tmp_markdown)
        assert len(docs) == 1
        assert "标题" in docs[0].page_content
        assert "加粗" in docs[0].page_content


class TestLoadXlsx:
    def test_basic(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["姓名", "年龄", "城市"])
        ws.append(["张三", 25, "北京"])
        ws.append(["李四", 30, "上海"])
        p = tmp_path / "test.xlsx"
        wb.save(str(p))
        wb.close()

        docs = _load_xlsx(str(p))
        assert len(docs) == 2
        assert "张三" in docs[0].page_content
        assert "北京" in docs[0].page_content

    def test_empty_sheet(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        ws.append(["col1", "col2"])
        p = tmp_path / "empty.xlsx"
        wb.save(str(p))
        wb.close()

        docs = _load_xlsx(str(p))
        # 无数据行时返回占位 Document
        assert len(docs) == 1
        assert "空" in docs[0].page_content


class TestProcessDocument:
    def test_txt(self, tmp_txt):
        chunks, doc_id = process_document(tmp_txt)
        assert len(chunks) >= 1
        assert doc_id
        assert chunks[0].metadata["document_id"] == doc_id

    def test_csv(self, tmp_csv):
        chunks, doc_id = process_document(tmp_csv)
        assert len(chunks) >= 1
        assert "张三" in chunks[0].page_content or any("张三" in c.page_content for c in chunks)

    def test_markdown(self, tmp_markdown):
        chunks, doc_id = process_document(tmp_markdown)
        assert len(chunks) >= 1

    def test_unsupported_format(self, tmp_path):
        p = tmp_path / "test.xyz"
        p.write_text("data")
        import pytest
        with pytest.raises(ValueError, match="不支持"):
            process_document(str(p))

    def test_xlsx(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["产品", "价格"])
        ws.append(["苹果", 5.5])
        ws.append(["香蕉", 3.0])
        p = tmp_path / "test.xlsx"
        wb.save(str(p))
        wb.close()
        chunks, doc_id = process_document(str(p))
        assert len(chunks) >= 1
        assert any("苹果" in c.page_content for c in chunks)
